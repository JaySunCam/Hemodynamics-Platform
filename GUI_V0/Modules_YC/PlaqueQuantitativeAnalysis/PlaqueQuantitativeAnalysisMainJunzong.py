# -*- coding: UTF-8 -*-
'''
@Project ：GUI 
@File    ：PlaqueQuantitativeAnalysis.py
@IDE     ：PyCharm 
@Author  ：YangChen's Piggy
@Date    ：2021/8/15 14:13 
'''
########################################################################################################################
import sys
import os
import time
#nowStr = datetime.now().strftime("%d/%b/%y - %H:%M:%S")
os.chdir(os.getcwd())
sys.path.insert(0, '../../Functions_YC')
sys.path.insert(0, '../../Functions_JZ')
# Import functions
import Save_Load_File
from FileDisposing import *
########################################################################################################################
# Standard library
import cv2
import copy
import json
import numpy
import openpyxl
import time

import skimage.measure
import skimage.segmentation
import skimage.morphology
from skimage.draw import line
import skimage.segmentation
import skimage.feature

from datetime import datetime
########################################################################################################################
# Standard libs

from scipy.spatial import KDTree
from scipy.interpolate import splprep, splev
from scipy import ndimage

import math
import matplotlib.pyplot as plt
import SimpleITK
################################################################################
def WaterShadeMsk(dataMat, Threshold = 0):
    image = dataMat > Threshold
    ############################################################################
    # Generate the markers as local maxima of the distance to the background
    # Now we want to separate the two objects in image
    # Generate the markers as local maxima of the distance
    # to the background
    distance = ndimage.distance_transform_edt(image)
    local_maxi = skimage.feature.peak_local_max(
        distance, indices=False, footprint=numpy.ones((3, 3)), labels=image)
    markers = skimage.measure.label(local_maxi)
    labels_ws = skimage.segmentation.watershed(-distance, markers, mask=image)

    markers[~image] = -1
    labels_rw = skimage.segmentation.random_walker(image, markers)

    return labels_rw

########################################################################################################################
def fillhole(Mask):
    Mask_fillhole = []
    for slicenum in range(0, len(Mask)):
        MskSliceArryfillhole = skimage.morphology.closing(Mask[slicenum],
                                                  selem=None)  # dilation  and  erosion
        Mask_fillhole.append(MskSliceArryfillhole)
    return Mask_fillhole

def RemoveSmallObj(Mask, size = 3):
    Mask_RSO = skimage.morphology.remove_small_objects(Mask, size)
    return Mask_RSO
########################################################################################################################
def Linedefine(point, angle):
     '''
     point - Tuple (x, y)
     angle - Angle you want your end point at in degrees.
     return
     point , slope ,intercept
     '''
     # unpack the first point
     x, y = point
     slope_k = math.tan(math.radians(angle))
     if angle not in [90,270]:
         intercept_b = y - slope_k*x
     else:
         slope_k = float("inf")
         intercept_b = float("inf")
         print("the intercept_b is infinite")
     return slope_k, intercept_b
########################################################################################################################
def ComponentContourExtraction(MskSliceArry , ObjectRegionLabel,Connectivity = False, arrayDataType = numpy.uint8):
    '''
         Arguements:
         MskSliceArry - origin image
         ObjectLabel - a list contain labels of the object region
         Closed_contour - specify whether shape is a closed contour (if passed True), or just a curve

         return:
         countour point , area , countour width
         '''
    # ********************** contours Property ********************************
    contourInfo = {}
    contourInfo['label'] = []
    contourInfo['area'] = []
    contourInfo['centroid'] = []
    contourInfo['height'] = []
    contourInfo['width'] = []
    contourInfo['coordinates'] = []

    # ************************binary image***************************************************
    MaskOnes = numpy.zeros_like(MskSliceArry)
    if numpy.sum(MskSliceArry):
        LabellList = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        ObjectRegionLabel.sort()
        LumenMaskOnes = numpy.zeros_like(MskSliceArry) #initialize LumenMaskOnes's size
        LumenMaskOnes[MskSliceArry == 1] = 1
        try:
            for labelcont in range(len(ObjectRegionLabel)):
                if ObjectRegionLabel[labelcont] not in LabellList:
                    MaskOnes[
                        MskSliceArry == ObjectRegionLabel[labelcont]] = 0
                else:
                    MaskOnes[
                        MskSliceArry == ObjectRegionLabel[labelcont]] = 1
        except:
            print(
                "Component Contour Extraction Error,Please make sure the "
                "label value is less than 10 ")

        # ******************* contours finding*******************
        MaskOnes_binary_image = MaskOnes.astype(arrayDataType)

        MaskOnes_contours, _ = cv2.findContours(MaskOnes_binary_image,
                                                cv2.RETR_TREE,
                                                cv2.CHAIN_APPROX_SIMPLE)
        # ************* filter the smaller  area *******************
        if len(MaskOnes_contours):
            if Connectivity == True:
                arealist = []
                for Mccont in range(len(MaskOnes_contours)):
                    contourarea = cv2.contourArea(MaskOnes_contours[Mccont])
                    arealist.append(contourarea)
                max_index = arealist.index(max(arealist))

                rect = cv2.minAreaRect(MaskOnes_contours[max_index])
                rect_label = 1
                rect_centroid = rect[0]
                rect_width = rect[1][0]
                rect_height = rect[1][1]
                rect_area = cv2.contourArea(MaskOnes_contours[max_index])

                contourInfo['label'].append(rect_label)
                contourInfo['area'].append(rect_area)
                contourInfo['centroid'].append(rect_centroid)
                contourInfo['height'].append(rect_height)
                contourInfo['width'].append(rect_width)
                contourInfo['coordinates'].append(MaskOnes_contours[max_index])
            else:
                for cont in range(len(MaskOnes_contours)):
                    rect = cv2.minAreaRect(MaskOnes_contours[cont])
                    rect_label = cont
                    rect_centroid = rect[0]
                    rect_width = rect[1][0]
                    rect_height = rect[1][1]
                    rect_area = cv2.contourArea(MaskOnes_contours[cont])

                    contourInfo['label'].append(rect_label)
                    contourInfo['area'].append(rect_area)
                    contourInfo['centroid'].append(rect_centroid)
                    contourInfo['height'].append(rect_height)
                    contourInfo['width'].append(rect_width)
                    contourInfo['coordinates'].append(MaskOnes_contours[cont])
                # else:
                #     #print("Warning: Current layer has no correspondent component!!")
    # else:
    #     print("Warning: Current layer is blank!!")
    return MaskOnes, contourInfo

def ComponentAreaCount(MskSliceArry , ObjectRegionLabel):
    # ************************binary image***************************************************
    PixelArea = 0
    tmpMask = copy.deepcopy(MskSliceArry)
    if numpy.sum(MskSliceArry):
        for ele in ObjectRegionLabel:
            PixelArea += numpy.sum(tmpMask == ele)
    PixelArea = int(PixelArea)
    return PixelArea
###############################################################################
def splinefitting(contour):
    smoothened = []
    #print("###################contour.T", type(contour), contour.T)
    x, y = contour.T
    #Convert from numpy arrays to normal arrays
    # x = x.tolist()[0]
    # y = y.tolist()[0]
    tck, u = splprep([x, y], u=None, s=2, per=True)
    #
    u_new = numpy.linspace(u.min(), u.max(), num=50)
    #
    x_new, y_new = splev(u_new, tck, der=0)
    # Convert it back to numpy format for opencv to be able to display it
    res_array = [[[int(i[0]), int(i[1])]] for i in zip(x_new, y_new)]
    smoothened.append(numpy.asarray(res_array, dtype=numpy.int16))
    return smoothened
    ##########################################################################
def IntersectionLineContour(contour, centroid, anglesize = 30):
    IntersectionPoints = []
    # ****************the boundingRect  of the contour*********************************
    x, y, w, h = cv2.boundingRect(contour)  #the contour's bounding rectinfo x,y,w,h
    #print("boundingrectinfo",x,y,w,h)
    # ****************the intersect points  of line and boundingRect********************
    for angle in range(0, 360, anglesize):
        infinite_angle = [90, 270]

        if angle not in infinite_angle:
            slope_k, intercept_b = Linedefine(centroid, angle)
            pa, pb = (x, int(slope_k * x + intercept_b)), \
                     ((x + w), int(slope_k * (x + w) + intercept_b))
        else:
            pa, pb = (int(centroid[0]), y), (int(centroid[0]), y + h)

        oneAngelIntersectionPoints = []
        tmponeAngelIntersectionPoints = []
        onepixeldist = 1
        for pt in zip(*line(*pa, *pb)):
            dist = cv2.pointPolygonTest(contour, (int(pt[0]), int(pt[1])), True)
            if cv2.pointPolygonTest(contour, (int(pt[0]), int(pt[1])), False) == 0:  # 若点在轮廓上
                # cv2.circle(contour, pt, 2, (0, 0, 255), 2)
                #print("pt on the edge", pt)
                tmponeAngelIntersectionPoints.append(pt)
            elif cv2.pointPolygonTest(contour, (int(pt[0]), int(pt[1])), False) > 0:
                if dist <= 1:
                    #print("pt near the edge", pt)
                    tmponeAngelIntersectionPoints.append(pt)
        try:
            oneAngelIntersectionPoints.append(tmponeAngelIntersectionPoints[0])
            oneAngelIntersectionPoints.append(tmponeAngelIntersectionPoints[-1])

            IntersectionPoints.append(oneAngelIntersectionPoints)
        except:
            print("############################### Functiong :Intersection Points ")
    return IntersectionPoints
########################################################################################################################
def KNNMatching(InitMat, ObjctMat, IniStartColumnNum, IniEndColumnNum, ObjStartColumnNum, ObjEndColumnNum,IniReturnColumnNum):
    ############################################find the nearest node###################################################
    # ********************************* the nearst node index  ********************************************************#
    MatchedRows_num = []
    # **************************** ***** the nearst distance  **********************************************************#
    min_distance = []
    # ********************************* the  Matched Rows datas in ObjctMat ******************************************#
    MatchedRows = []
    # ********************************* one culumn of  Matched Rows datas in ObjctMat  *******************************#
    MatchedRows_1stColumn = []
    # ******************* the initial datas including all information like pressure coordinates elementid  ************#
    tree = KDTree(InitMat[:,IniStartColumnNum : IniEndColumnNum+1])  #ObjctMat[:,0]is the NodeID of wall elements
    for i in range(len(ObjctMat)):
        dist, InitMatRowsQueryed = tree.query(ObjctMat[i,ObjStartColumnNum : ObjEndColumnNum+1], k=1)
        MatchedRows_num.append(InitMatRowsQueryed)
        MatchedRows.append(InitMat[InitMatRowsQueryed,:])
        MatchedRows_1stColumn.append(InitMat[InitMatRowsQueryed, IniReturnColumnNum]) #return nodeid
        min_distance.append(dist)
    return MatchedRows_num, min_distance, MatchedRows, MatchedRows_1stColumn
########################################################################################################################
def skeleton_demo(binaryimage):
     skeleton0 = skimage.morphology.skeletonize(binaryimage)
     skeleton = skeleton0.astype(numpy.uint8) * 255
     # cv2.imshow("skeleton", skeleton)
     # cv2.waitKey(0)
     # cv2.destroyAllWindows()
     return skeleton
########################################################################################################################
def removeSelectedContour(image_src, CountourAreaOrder):
    contours, hierarchy = cv2.findContours(image_src, cv2.RETR_LIST,
                                                  cv2.CHAIN_APPROX_SIMPLE)
    mask = numpy.zeros(image_src.shape, numpy.uint8)
    largest_areas = sorted(contours, key=cv2.contourArea)
    cv2.drawContours(mask, [largest_areas[-(CountourAreaOrder)]], 0,(255,255,255),1)
    removed = cv2.subtract(image_src, mask)
    # cv2.imshow("removedskeleton", removed)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    return removed
########################################################################################################################
def Cal_kb_linear_fitline(loc):
    output = cv2.fitLine(loc, cv2.DIST_L2, 0, 0.01, 0.01)
    k = output[1] / output[0]
    b = output[3] - k * output[2]
    return k,b

def curvefitting(x_arry, y_arry, order):
    curveParameterArry = numpy.polyfit(x_arry, y_arry, order)
    curve = numpy.poly1d(numpy.polyfit(x_arry,
                                   y_arry,
                                   order))  # poly1d :get the

    plt.scatter(x_arry, y_arry)
    plt.plot(x_arry, curve(x_arry), color='g')
    print("curve", curve)
    return curve


########################################################################################################################
# app = QApplication([])
# MainWindow = QMainWindow()
# FileDialog = QFileDialog(MainWindow)
class PlaqueQuantitativAnalysis:
    """
    Description:
    Quantitative analysis of plaque.
    """
    def __init__(self,
                 InputFilePath = "", OutputFilePath = "",xlsxSavePath = ""):
        #workpath
        self.spacing = None
        self.spacing_x = None
        self.spacing_y = None
        self.spacing_z = None
        self.InputFilePath = InputFilePath
        self.OutputFilePath = OutputFilePath
        self.xlsxSavePath = xlsxSavePath
        self.IntDataType = None  #numpy.uint8
        self.FloatDataType = None  #numpy.float64
        # initial definition

        self.outputData = None
        self.outputDataRemoveBifurcation = None
        self.outputData_cont = None
        self.outputDatafillhole = None
        self.outputDataOnes = None
        self.CTA = None
        self.inMsk = None
        self.BranchNumsList = None
        self.BranchSliceGroup_ResetLabelDict = None
        self.slicesNum = None
        self.BranchGroupMask = None
        self.LumenGroup = None
        self.BranchGroup = None
        self.Lumen_contours = None
        self.subLumen_contours = None
        self.contourspline = None
        self.sliceCCentroid = None
        self.sliceIntersectPointsInfo = None
        #wall property
        self.wallthicknessInfo = None
        self.wallAreaInfo = None
        self.AverWallThickness =None
        self.WallThicknessRange =None
        self.singleAnglewallthickness = None
        self.slicew = None
        self.PlaqueBurdenInfo = None
        self.PlaqueBurden = None
        self.WallEqDiameterInfo = None
        self.LumenEqDiameterInfo = None

        #component volumn and area
        self.TotalCalcificationMaskArea = None
        self.TotalCalcificationArea = None
        self.TotalCalcificationVolumn = None
        self.TotalCalcificationVolumnRatio = None

        self.TotalLipidMaskArea = None
        self.TotalLipidArea = None
        self.TotalLipidVolumn = None
        self.TotalLipidVolumnRatio = None

        self.TotalPlaqueMaskArea = None
        self.TotalPlaqueArea = None
        self.TotalPlaqueVolumn = None

        self.TotalHemorrhageMaskArea = None
        self.TotalHemorrhageArea = None
        self.TotalHemorrhageVolumn = None
        self.TotalHemorrhageVolumnRatio = None

        self.TotalFibrousMaskArea = None
        self.TotalEnhancementMaskArea = None
        self.TotalFibrousArea = None
        self.TotalFibrousVolumn = None
        self.TotalFibrousVolumnRatio = None

        self.TotalLumenArea = None
        self.AverLumenArea = None
        self.LumenAreaRange = None
        self.WallAreaInfo = None
        #component height
        self.PlaqueHeight = None
        self.BranchGroupFibrousIntegrity = None
        self.BranchGroupSliceLumenMaskAreadict = None
        self.BranchGroupSlicePlaqueMaskAreadict = None
        self.BranchGroupslicePlaqueAreaBooldict = None
        self.BranchGroupslicBifurcationPositionlist = None
        self.stenosisLumenEqDiameter = None
        self.stenosisPositionInfo = None
        self.BranchGroupreferenceWallEquivalentDiameterIndex = None
        self.BranchGroupreferenceWallEquivalentDiameter = None
        self.BranchGroupStenosisDegree = None
        self.StenosisDegreeList = None
        self.stenosisdegree = None
        self.ReconstructionIndex =None
        self.ReconstructionIndexMin =None
        self.walleccentricity = None
        self.BranchGroupReconstructionIndex = None
        self.PlaqueAnalysisInfo = None
        self.blankLayer = None
        self.bifurcation = None
        self.nonblanklayerIdlist = None
        self.BranchNumsListremove0 = None

        self.rtrnInfo = {}
        self.rtrnInfo["error"] = False
        self.rtrnInfo["errorMessage"] = ""
        self.rtrnInfo["warningMessage"] = ""
        self.rtrnInfo["processTime"] = None
        self.rtrnInfo["processTimeMessage"] = None
        self.rtrnInfo["ControlPoints"] = None
        self.rtrnInfo["sortArray"] = None
        self.rtrnInfo["message"] = ""
        self.rtrnInfo["processTime"] = None

        self.WebrtrnInfo = {}
        self.WebrtrnInfo["error"] = False
        self.WebrtrnInfo["errorMessage"] = ""
        self.WebrtrnInfo["warningMessage"] = ""
        self.WebrtrnInfo["processTime"] = None
        self.WebrtrnInfo["processTimeMessage"] = None
        self.WebrtrnInfo["ControlPoints"] = None
        self.WebrtrnInfo["sortArray"] = None
        self.WebrtrnInfo["message"] = ""
        self.WebrtrnInfo["processTime"] = None
        self.WebrtrnInfo["log"] = ""

        # rtrnInfo["processTime"] = stpT - strtT
        print("#################################################################\n"
              "########### the default label of each component #################\n"
              "1 - Lumen\n"
              "2 - Healthy Wall + other/wall\n"
              "3 - Calcification\n"
              "4 - Lipid\n"
              "5 - Loose matrix(fibrous tisue)\n"
              "6 - Hemorrhage - Fresh\n"
              "7 - Hemorrhage - Recent\n"
              "8 - non-evaluated (NE)\n"
              "#################################################################\n")

    def LoadData(self):
        self.PlaqueAnalysisInfo = {}
        #load data
        self.inMsk = Save_Load_File.OpenLoadNIFTI(
            GUI = False,
            filePath = self.InputFilePath
        )
        # update display data
        self.outputDataInit = copy.deepcopy(self.inMsk.OriData)
        self.outputData = self.outputDataInit
        self.IntDataType = self.inMsk.OriData.dtype
        try:
            self.outputDataOnes = numpy.array(
                (self.outputData != 0) * 1,
                dtype=self.inMsk.OriData.dtype.type
            )
        except:
            self.outputDataOnes = numpy.array(
                (self.outputData != 0) * 1,
                dtype=numpy.int16
            )
        itkImag = SimpleITK.ReadImage(self.InputFilePath)
        # get voxel spacing (for 3-D image)

        self.spacing = itkImag.GetSpacing()
        self.spacing_x = self.spacing[0]
        self.spacing_y = self.spacing[1]
        self.spacing_z = self.spacing[2]
        self.rtrnInfo["message"] += "Load data success! "
    def DataBifurcationRemove(self):
        self.WebrtrnInfo["warningMessage"] = ""
        self.WebrtrnInfo["log"] += \
            "   spacing_x: " + str(self.spacing_x) + \
            "   spacing_y: " + str(self.spacing_y) + \
            "   spacing_z: " + str(self.spacing_z)
        self.slicBifurcationlayers = []
        for slicenum in range(0, len(self.outputData)):
            if numpy.sum(self.outputData[slicenum]):
                MskSliceArry = self.outputData[slicenum]
                Wall_binary_image, WallContourInfo = \
                    ComponentContourExtraction(MskSliceArry,
                                               [1,2,3,4,5,6,7,8,9])
                Lumen_binary_imageRef, LumenContourInfoRef = \
                    ComponentContourExtraction(MskSliceArry,
                                               [1])
                SliceBranchNumsRef = len(LumenContourInfoRef['area'])
                if SliceBranchNumsRef > 2:
                    self.WebrtrnInfo["errorMessage"] += \
                         " layer " + str(slicenum + 1) + " has more than 2 lumens! "
                    self.PlaqueAnalysisInfo["errorMessage"] = self.WebrtrnInfo["errorMessage"]
                    self.JsonFileOutput()
                    sys.exit(0)
                elif SliceBranchNumsRef == 2:
                    if len(WallContourInfo['area']) ==1:
                        self.slicBifurcationlayers.append(slicenum)
        print("self.slicBifurcationlayers",self.slicBifurcationlayers)
        tmpoutputDataRemoveBifurcation = copy.deepcopy(self.outputData)
        self.outputDataRemoveBifurcation = numpy.delete(tmpoutputDataRemoveBifurcation, self.slicBifurcationlayers, axis=0)
################################Normal##################################
    def LumenBinary(self):
        # split the slice with two lumens
        self.BranchNumsList = []
        self.LumenGroup = []
        self.blankLayer = []
        # *******************get lumen binary image\ branchnums\blank layer num**********************
        for slicenum in range(0, len(self.outputData_cont)):
            # ********************fill the incorrect hole **********************
            if numpy.sum(self.outputData_cont[slicenum]):   # if the current layer has date
                MskSliceArry = self.outputData_cont[slicenum]
                Lumen_binary_image, LumenContourInfo = \
                    ComponentContourExtraction(MskSliceArry,
                                               [1])
                SliceBranchNums = len(LumenContourInfo['area'])
                self.LumenGroup.append(Lumen_binary_image)
                self.BranchNumsList.append(SliceBranchNums)
            else:
                self.BranchNumsList.append(0)
                self.blankLayer.append(slicenum)
                self.rtrnInfo["warningMessage"] += \
                    "***********************************\n" + \
                    " layer " + str(slicenum + 1) + " is blank\n"

        if sum(self.BranchNumsList) == 0:
            self.WebrtrnInfo["errorMessage"] += "The current Mask is blank! "
            self.PlaqueAnalysisInfo["errorMessage"] = \
                self.WebrtrnInfo["errorMessage"]
            self.JsonFileOutput()
            sys.exit(0)

    def NonBlanklayerIdGet(self):
        self.nonblanklayerIdlist = []
        self.BranchNumsListremove0 = []
        for cont in range(len(self.BranchNumsList)):
            if self.BranchNumsList[cont] != 0:
                self.BranchNumsListremove0.append(self.BranchNumsList[cont])
                self.nonblanklayerIdlist.append(cont)
        if list(set(self.BranchNumsListremove0)) == [2]:
            self.WebrtrnInfo["warningMessage"] += "The current Mask has no cervical artery!"
            self.PlaqueAnalysisInfo["warningMessage"] = \
                self.WebrtrnInfo["warningMessage"]
            self.JsonFileOutput()

    def BifurcationGet(self):
        # ******************branch bifurcation
        for cont in range(len(self.BranchNumsListremove0)):
            if self.BranchNumsListremove0[cont] == 2:
                self.bifurcation = cont
                break
            else:
                self.bifurcation = cont + 1
        print("self.bifurcation ", self.bifurcation +1)
        if  self.bifurcation == 0:
            self.WebrtrnInfo[
                "warningMessage"] = "The current Mask has no cervical artery"
            self.JsonFileOutput()
        self.rtrnInfo["message"] += \
            "The bufurcation positon of the carotid is %s\n" % (str(self.bifurcation + 1))

    def BranchSplited(self):  ##do not need branch spliting
        self.BranchSliceGroup_ResetLabelDict = {}
        self.BranchNums = sum(list(set(self.BranchNumsList))) #branch numbers
        for cont in range(len(self.nonblanklayerIdlist)):
            slicenum = self.nonblanklayerIdlist[cont]
            MskSliceArry = self.outputData_cont[slicenum]
            MaskAllSliceOnes, MaskAllSliceContourInfo = \
                ComponentContourExtraction(MskSliceArry,
                                           [1, 2, 3, 4, 5, 6, 7, 8, 9],
                                           Connectivity = False)
            MaskRemoveLumenSliceOnes, MaskRemoveLumenSliceContourInfo = \
                ComponentContourExtraction(MskSliceArry,
                                           [2, 3, 4, 5, 6, 7, 8, 9],
                                           Connectivity = False)
            #************************** reset label to 1 or 2******************
            if self.BranchNumsList[slicenum] == 1:
                area_splited_ResetLabel = copy.deepcopy(MskSliceArry)
                area_splited_ResetLabel[area_splited_ResetLabel != 0] = 1

                self.BranchSliceGroup_ResetLabelDict['layer' + str(cont)] = \
                    area_splited_ResetLabel

            elif self.BranchNumsList[slicenum] > 1:
                # ***************area pre-cutting  skeleton method**************
                MaskAllSliceOnesAstype = MaskAllSliceOnes.astype(numpy.uint8) * 255
                # cv2.imshow("MaskAllSliceOnesAstype", MaskAllSliceOnesAstype)
                # cv2.waitKey(0)
                # cv2.destroyAllWindows()
                MaskAllSliceOnesContours, _ = cv2.findContours(MaskAllSliceOnesAstype,
                                                        cv2.RETR_TREE,
                                                        cv2.CHAIN_APPROX_SIMPLE)
                tmpSliceContourFilled = numpy.zeros_like(MskSliceArry)

                if len(MaskAllSliceOnesContours) == self.BranchNumsList[slicenum]:
                    for i in range(len(MaskAllSliceOnesContours)):
                        cv2.drawContours(tmpSliceContourFilled, MaskAllSliceOnesContours, i, i+1,
                                         -1)
                    area_splited_ResetLabel = tmpSliceContourFilled

                else:
                    wallskeleton = skeleton_demo(MaskRemoveLumenSliceOnes)
                    initCutOffLine = removeSelectedContour(wallskeleton, 1) #remove the outline

                    # 取出矩阵中的非零元素的坐标
                    initCutOffLineArry = numpy.mat(initCutOffLine)

                    nonzeroCoodinateTuple = initCutOffLineArry.nonzero()
                    nonzeroCoodinateArry = numpy.asarray(nonzeroCoodinateTuple).T

                    if len(nonzeroCoodinateArry):
                        # ************fit line**********
                        k, b = Cal_kb_linear_fitline(nonzeroCoodinateArry)

                        # ************split region by line**********
                        MskSliceArry_h, MskSliceArry_w = MskSliceArry.shape[:2]
                        # print("MskSliceArry_h", MskSliceArry_h,
                        #       "MskSliceArry_w",
                        #       MskSliceArry_w)
                        MaskAllSliceROI = numpy.zeros_like(MskSliceArry)

                        # print("k", k, "b", b, "MaskAllSliceROI",
                        #       MaskAllSliceROI.shape)
                    else:
                        self.WebrtrnInfo["errorMessage"] += "Can not generate split line in " \
                                              "layer %s, please make sure " \
                                              "there is no single point " \
                                              "connection!" % (
                            str(slicenum + 1))
                        self.PlaqueAnalysisInfo["errorMessage"] = self.WebrtrnInfo["errorMessage"]
                        self.JsonFileOutput()
                        sys.exit(0)

                    for h_cont in range(int(MskSliceArry_h)):
                        for w_cont in range(int(MskSliceArry_w)):
                            # #************split region by line**********
                            if ((h_cont * k + b) <= w_cont):
                                # MaskAllSliceROI[0, 0] = 1
                                MaskAllSliceROI[h_cont, w_cont] = 1
                            else:
                                # MaskAllSliceROI[0, 0] = 0
                                MaskAllSliceROI[h_cont, w_cont] = 0
                            # ************split region by curve*********
                            # if curve(w_cont) <= h_cont:
                            #     MaskAllSliceROI[w_cont, h_cont] = 1
                            # else:
                            #     MaskAllSliceROI[w_cont, h_cont] = 0
                    # plt.figure().suptitle('MaskAllSliceROI')
                    # plt.imshow(MaskAllSliceROI + MskSliceArry )
                    area_splited_ResetLabel = 2 * numpy.multiply(MaskAllSliceOnes, MaskAllSliceROI) + \
                                              numpy.multiply(MaskAllSliceOnes, 1 - MaskAllSliceROI)
            self.BranchSliceGroup_ResetLabelDict['layer' + str(cont)] = \
                area_splited_ResetLabel
        print("self.slicBifurcationlayers", self.slicBifurcationlayers)

    def BranchGrouping(self):
        self.BranchGroup = {}
        BranchGroupMask = {}

        # **************************branch grouping*****************************
        # *************************init branch group****************************
        for BNcont in range(self.BranchNums):
            BranchGroupMask['branch' + str(BNcont + 1)] = []
            self.BranchGroup['branch' + str(BNcont + 1)] = []
        # *************************branch grouping****************************
        for cont in range(len(self.BranchSliceGroup_ResetLabelDict)):
            Mask1 = numpy.zeros_like(
                self.BranchSliceGroup_ResetLabelDict['layer0'])
            Mask2 = numpy.zeros_like(
                self.BranchSliceGroup_ResetLabelDict['layer0'])
            if cont < self.bifurcation :
                BranchGroupMask['branch' + str(1)].append(
                    self.BranchSliceGroup_ResetLabelDict[
                        'layer' + str(cont)])
                self.BranchGroup['branch' + str(1)].append(
                    self.outputData_cont[self.nonblanklayerIdlist[cont]])
            else:
                if cont < len(self.BranchSliceGroup_ResetLabelDict) -1:
                    twolayer_Multiple = numpy.multiply(
                        self.BranchSliceGroup_ResetLabelDict[
                            'layer' + str(cont)],
                        self.BranchSliceGroup_ResetLabelDict[
                            'layer' + str(cont + 1)])

                    # twolayer_Multiple = twolayer_Multiple.astype(numpy.uint8) * 255
                    # cv2.imshow("twolayer_Multiple",
                    # twolayer_Multiple)
                    # cv2.waitKey(0)
                    # cv2.destroyAllWindows()
                    # plt.figure().suptitle(str(cont))
                    # plt.imshow(twolayer_Multiple)


                    twolayer_MultipleRremove0 = twolayer_Multiple[
                        twolayer_Multiple != 0].flatten()
                    Ele_repeatnum = numpy.bincount(twolayer_MultipleRremove0)
                    # print("Ele_repeatnum", Ele_repeatnum, "current layer",
                    # cont)
                    try:
                        Ele_repeatnum_largest = numpy.argmax(Ele_repeatnum)
                    except:
                        self.WebrtrnInfo["errorMessage"] += "line " + str(
                            sys._getframe().f_lineno) + ": There is no intersection part between " \
                                                        "relative layer " + str(
                            cont + 1) + " and layer " + str(cont + 2) + ", please make sure those two layer is continuous!"
                        self.PlaqueAnalysisInfo["errorMessage"] = \
                        self.WebrtrnInfo["errorMessage"]
                        self.JsonFileOutput()
                    twolayer_MultipleNonrepeatingElementRemove0 = numpy.unique(
                        twolayer_MultipleRremove0).tolist()

                    firstLayerNonrepeatingElement = \
                        numpy.unique(self.BranchSliceGroup_ResetLabelDict[
                                         'layer' + str(cont)])
                    firstLayerNonrepeatingElementRemove0 = \
                        firstLayerNonrepeatingElement[
                            firstLayerNonrepeatingElement != 0].tolist()

                    secondLayerNonrepeatingElement = \
                        numpy.unique(self.BranchSliceGroup_ResetLabelDict[
                                         'layer' + str(cont + 1)])
                    secondLayerNonrepeatingElementRemove0 = \
                        secondLayerNonrepeatingElement[
                            secondLayerNonrepeatingElement != 0].tolist()

                    tmps1 = numpy.zeros_like(
                        self.BranchSliceGroup_ResetLabelDict[
                            'layer' + str(cont + 1)])
                    tmps2 = numpy.zeros_like(
                        self.BranchSliceGroup_ResetLabelDict[
                            'layer' + str(cont + 1)])

                    # BranchSliceGroup_ResetLabelDict
                    if len(firstLayerNonrepeatingElementRemove0) == len(
                            secondLayerNonrepeatingElementRemove0) == 2:

                        if twolayer_MultipleNonrepeatingElementRemove0 in [
                            [1, 2], [2, 4], [2]]:
                            tmps1[self.BranchSliceGroup_ResetLabelDict[
                                      'layer' + str(cont + 1)] == 2] = 1
                            tmps2[self.BranchSliceGroup_ResetLabelDict[
                                      'layer' + str(cont + 1)] == 1] = 2
                            subtmp = tmps1 + tmps2
                            self.BranchSliceGroup_ResetLabelDict[
                                'layer' + str(cont + 1)] = subtmp

                    elif len(firstLayerNonrepeatingElementRemove0) == len(
                            secondLayerNonrepeatingElementRemove0) == 1:

                        if twolayer_MultipleNonrepeatingElementRemove0 == [2]:
                            tmps1[self.BranchSliceGroup_ResetLabelDict[
                                      'layer' + str(cont + 1)] == 2] = 1
                            tmps2[self.BranchSliceGroup_ResetLabelDict[
                                      'layer' + str(cont + 1)] == 1] = 2
                            subtmp = tmps1 + tmps2
                            self.BranchSliceGroup_ResetLabelDict[
                                'layer' + str(cont + 1)] = subtmp

                    elif len(firstLayerNonrepeatingElementRemove0) != len(
                            secondLayerNonrepeatingElementRemove0):

                        if twolayer_MultipleNonrepeatingElementRemove0 == [2]:
                            tmps1[self.BranchSliceGroup_ResetLabelDict[
                                      'layer' + str(cont + 1)] == 2] = 1
                            tmps2[self.BranchSliceGroup_ResetLabelDict[
                                      'layer' + str(cont + 1)] == 1] = 2
                            subtmp = tmps1 + tmps2
                            self.BranchSliceGroup_ResetLabelDict[
                                'layer' + str(cont + 1)] = subtmp

                        elif twolayer_MultipleNonrepeatingElementRemove0 in [
                            [1, 2], [2, 4]]:
                            if Ele_repeatnum_largest == 2:
                                tmps1[self.BranchSliceGroup_ResetLabelDict[
                                          'layer' + str(cont + 1)] == 2] = 1
                                tmps2[self.BranchSliceGroup_ResetLabelDict[
                                          'layer' + str(cont + 1)] == 1] = 2
                                subtmp = tmps1 + tmps2
                                self.BranchSliceGroup_ResetLabelDict[
                                    'layer' + str(cont + 1)] = subtmp

                # save the current layer to  BranchGroupMask
                Mask1[self.BranchSliceGroup_ResetLabelDict[
                          'layer' + str(cont)] == 1] = 1
                # plt.figure().suptitle('Mask1  ' + str(cont))
                # plt.imshow(Mask1)

                Mask2[self.BranchSliceGroup_ResetLabelDict[
                          'layer' + str(cont)] == 2] = 2

                # plt.figure().suptitle('Mask2' + str(cont))
                # plt.imshow(Mask2)
                BranchGroupMask['branch' + str(self.BranchNums - 1)].append(
                    Mask1)  # BranchGroupMask is a logical matrix here
                BranchGroupMask['branch' + str(self.BranchNums - 0)].append(Mask2)

                MatrixSplited1 = numpy.multiply(
                    self.outputData_cont[self.nonblanklayerIdlist[cont]],
                    Mask1)

                MatrixSplited2 = numpy.multiply(
                    self.outputData_cont[self.nonblanklayerIdlist[cont]],
                    Mask2 * 0.5)

                if numpy.sum(MatrixSplited1):
                    self.BranchGroup['branch' + str(self.BranchNums - 1)].append(MatrixSplited1)######
                if numpy.sum(MatrixSplited2):
                    self.BranchGroup['branch' + str(self.BranchNums - 0)].append(MatrixSplited2)#####
            # plt.figure().suptitle('layer' + str(cont))
            # plt.imshow(self.BranchSliceGroup_ResetLabelDict['layer' + str(cont)])
        plt.show()

    def IntersectPoints(self):
        self.contourspline = []
        self.sliceCCentroid = {}
        self.sliceIntersectPointsInfo = {}
        self.PlaqueBurdenInfo ={}
        self.PlaqueBurden = 0
        slicePlaqueBurdenlist = []
        self.WallEqDiameterInfo = {}
        self.LumenEqDiameterInfo = {}
        self.WallAreaInfo = {}
        for BGcont in range(len(self.BranchGroup)):
            self.sliceCCentroid['branch' + str(BGcont + 1)] = []
        try:
            for BGcont in range(len(self.BranchGroup)):
                slicewallArealist = []
                sliceWallEqDiameterlist = []
                sliceLumenEqDiameterlist = []
                tmpbranch = self.BranchGroup['branch' + str(BGcont + 1)]
                for slicenum in range(len(self.BranchGroup['branch' + str(BGcont + 1)])):
                    current_subbranch = tmpbranch[slicenum]
                    # ************************** contour of wall inner and outer contour **********************************************
                    WallMaskOnes, WallcontourInfo = \
                        ComponentContourExtraction(current_subbranch,
                                                   [1, 2, 3, 4, 5, 6, 7, 8, 9],
                                                   Connectivity=False)

                    Lumen_binary_image, LumenContourInfo = \
                        ComponentContourExtraction(current_subbranch, [1],
                                                   Connectivity=False)

                    slicewallArea = ComponentAreaCount(current_subbranch, [1, 2, 3, 4, 5, 6, 7, 8, 9])
                    sliceWallEqDiameter = numpy.sqrt(4 * slicewallArea / numpy.pi)
                    slicewallArealist.append(slicewallArea)
                    sliceWallEqDiameterlist.append(sliceWallEqDiameter)

                    slicelumenArea = ComponentAreaCount(current_subbranch, [1])
                    sliceLumenEqDiameter = numpy.sqrt(4 * slicelumenArea / numpy.pi)
                    sliceLumenEqDiameterlist.append(sliceLumenEqDiameter)

                    slicePlaqueBurden = (slicewallArea - slicelumenArea) / slicewallArea
                    slicePlaqueBurdenlist.append(slicePlaqueBurden)

                    # ************************** the center of lumen inner  centroid ****************************************
                    Ccentroid_x = LumenContourInfo['centroid'][0][0]
                    Ccentroid_y = LumenContourInfo['centroid'][0][1]
                    Ccentroid = (Ccentroid_x, Ccentroid_y)
                    self.sliceCCentroid['branch' + str(BGcont + 1)].append(
                        Ccentroid)

                    # ***************************************intersect points*********************************
                    tmpIntersectionPoints = []
                    walltmpContour = WallcontourInfo['coordinates'][0]
                    wallIntersectionPoints = IntersectionLineContour(
                        walltmpContour, Ccentroid)

                    tmpIntersectionPoints.append(wallIntersectionPoints)
                    lumentmpContour = LumenContourInfo['coordinates'][0]
                    lumenIntersectionPoints = IntersectionLineContour(
                        lumentmpContour, Ccentroid)

                    tmpIntersectionPoints.append(lumenIntersectionPoints)
                    self.sliceIntersectPointsInfo['branch_' + str(BGcont + 1)
                        + "/layer_"+ str(slicenum)] = tmpIntersectionPoints

                # ***************************************PlaqueBurdenInfo*********************************
                self.PlaqueBurdenInfo[
                    'branch' + str(BGcont + 1)] = slicePlaqueBurdenlist
                # ***************************************wall area*********************************
                self.WallAreaInfo[
                    'branch' + str(BGcont + 1)] = slicewallArealist
                # ***************************************equivalent diameter*********************************
                self.WallEqDiameterInfo[
                    'branch' + str(BGcont + 1)] = sliceWallEqDiameterlist
                self.LumenEqDiameterInfo[
                    'branch' + str(BGcont + 1)] = sliceLumenEqDiameterlist
        except:
           print("There is a blank branch.")

        # ***************************************PlaqueBurden*********************************
        print("slicePlaqueBurdenlist", slicePlaqueBurdenlist)
        slicePlaqueBurdenlist.sort()
        self.PlaqueBurden = slicePlaqueBurdenlist[-1]
        print("PlaqueBurden",self.PlaqueBurden)


    def CarotidTissueParameters(self):
        self.wallAreaInfo = {}
        self.TotalPlaqueMaskArea = 0
        self.TotalPlaqueVolumn = 0
        self.TotalPlaqueVolumnRatio = 0

        ####
        self.TotalCalcificationMaskArea = 0
        self.TotalCalcificationArea = 0
        self.TotalCalcificationVolumn = 0
        self.TotalCalcificationVolumnRatio = 0

        self.TotalLipidMaskArea = 0
        self.TotalLipidArea = 0
        self.TotalLipidVolumn = 0
        self.TotalLipidVolumnRatio = 0

        self.TotalPlaqueMaskArea = 0

        self.TotalPlaqueArea = 0
        self.TotalPlaqueMaskArea_junzong =0
        self.TotalPlaqueArea_junzong = 0
        self.TotalPlaqueVolumn = 0

        self.TotalHemorrhageMaskArea = 0
        self.TotalHemorrhageArea = 0
        self.TotalHemorrhageVolumn = 0
        self.TotalHemorrhageVolumnRatio = 0

        self.TotalFibrousMaskArea = 0
        self.TotalEnhancementMaskArea = 0
        self.TotalFibrousArea = 0
        self.TotalFibrousVolumn = 0
        self.TotalFibrousVolumnRatio = 0

        self.TotalPlaqueArea = 0
        self.TotalPlaqueVolumn = 0
        self.TotalPlaqueVolumn_junzong = 0

        self.TotalFibrousMaskArea = 0
        self.TotalEnhancementMaskArea = 0
        self.TotalLumenMaskArea = 0
        self.PlaqueHeight = 0

        self.slicesNum = sum(self.BranchNumsList)
        self.BranchGroupSliceLumenMaskAreadict ={}
        self.BranchGroupslicePlaqueAreaBooldict = {}
        self.LumenAreaRange = []
        PlaqueMaskHeight =0
        for slicenum in range(0, len(self.outputData_cont)):
        # ************************plaque height ********************************
            MskSliceArry = self.outputData_cont[slicenum]
            PlaqueOnematrixsum = ComponentAreaCount(MskSliceArry, [3,4,5,6,7,8, 9])
            if PlaqueOnematrixsum:
                PlaqueMaskHeight += 1
        self.PlaqueHeight = PlaqueMaskHeight * self.spacing_z
        print("PlaqueHeight", self.PlaqueHeight)


        # **************** FibrousIntegrity *********************************
        self.BranchGroupFibrousIntegrity = {}
        self.branchGroupSliceAreaInfo = {}
        self.branchGroupSliceAreaMaskInfo = {}
        for BGcont in range(len(self.BranchGroup)):
            self.branchSliceAreaInfo = {}
            self.branchSliceAreaMaskInfo = {}
            sliceLumenMaskArealist = []
            slicePlaqueAreaBoollist = []
            print("######@@@@@@!!", len(self.BranchGroup['branch' + str(BGcont + 1)]))
            groupwallAreaInfo = {}
            for slicenum in range(len(self.BranchGroup['branch' + str(BGcont + 1)])):
                slicewallAreaInfo = {}
                slicewallAreaMaskInfo = {}
                currentslice = self.BranchGroup['branch' + str(BGcont + 1)][slicenum]

                sliceLumenArea = ComponentAreaCount(currentslice, [1])

                sliceLumenMaskArealist.append(sliceLumenArea)
                self.TotalLumenMaskArea += sliceLumenArea


                slicewallAreaMaskInfo["Lumen"] = sliceLumenArea
                slicewallAreaInfo["Lumen"] = sliceLumenArea * self.spacing_x * self.spacing_y

                slicewallAreaMaskInfo["Wall"] = ComponentAreaCount(currentslice, [1, 2, 3, 4, 5, 6, 7, 8, 9])
                slicewallAreaInfo["Wall"] = ComponentAreaCount(currentslice,[1, 2, 3, 4, 5,6, 7, 8, 9]) * self.spacing_x * self.spacing_y

                slicePlaque = ComponentAreaCount(currentslice, [3,4,5,6,7,8, 9])
                # ************************************where has the plaque ***********************************
                if slicePlaque >0:
                    slicePlaqueAreaBool = 1
                    slicePlaqueAreaRef = ComponentAreaCount(currentslice,
                                                            [2, 3, 4, 5, 6, 7,
                                                             8,9])
                    slicePlaqueArea_junzong = ComponentAreaCount(currentslice,
                                                            [3, 4, 5, 6, 7,
                                                             8, 9])
                else:
                    slicePlaqueAreaBool = 0
                    slicePlaqueAreaRef = 0
                    slicePlaqueArea_junzong = 0
                slicePlaqueAreaBoollist.append(slicePlaqueAreaBool)
                self.TotalPlaqueMaskArea += slicePlaqueAreaRef
                self.TotalPlaqueMaskArea_junzong += slicePlaqueArea_junzong

                sliceCalcificationArea = ComponentAreaCount(currentslice, [3])
                slicewallAreaMaskInfo["Calcification"] =  sliceCalcificationArea
                slicewallAreaInfo["Calcification"] =  sliceCalcificationArea * self.spacing_x * self.spacing_y
                self.TotalCalcificationMaskArea += sliceCalcificationArea

                sliceLipidArea = ComponentAreaCount(currentslice, [4])
                slicewallAreaMaskInfo["Lipid"] = sliceLipidArea
                slicewallAreaInfo["Lipid"] = sliceLipidArea * self.spacing_x * self.spacing_y
                self.TotalLipidMaskArea += sliceLipidArea

                sliceHemorrhageArea = ComponentAreaCount(currentslice, [6,7])
                slicewallAreaMaskInfo["Hemorrhage"] = sliceHemorrhageArea
                slicewallAreaInfo["Hemorrhage"] = sliceHemorrhageArea * self.spacing_x * self.spacing_y
                self.TotalHemorrhageMaskArea += sliceHemorrhageArea

                sliceFibrousOnes, sliceFibrouscontourInfo = ComponentContourExtraction(currentslice, [5])
                sliceFibrousArea = ComponentAreaCount(currentslice, [5])
                slicewallAreaMaskInfo["Fibrous"] = sliceFibrousArea
                slicewallAreaInfo["Fibrous"] = sliceFibrousArea * self.spacing_x * self.spacing_y
                self.TotalFibrousMaskArea += sliceFibrousArea

                sliceEnhancement, sliceEnhancementcontourInfo = \
                    ComponentContourExtraction(
                    currentslice, [9])
                sliceEnhancementArea = ComponentAreaCount(currentslice, [9])
                slicewallAreaMaskInfo["Enhancement"] = sliceEnhancementArea
                slicewallAreaInfo["Enhancement"] = sliceEnhancementArea * self.spacing_x * \
                                 self.spacing_y
                self.TotalEnhancementMaskArea += sliceEnhancementArea


                if len(sliceFibrouscontourInfo['area']) ==0 :
                    tmpFibrousIntegrity = "None"
                elif len(sliceFibrouscontourInfo['area']) >1:
                    tmpFibrousIntegrity = "true"
                else:
                    tmpFibrousIntegrity = "false"
                self.BranchGroupFibrousIntegrity['branch' + str(BGcont + 1) + ' layer' + str(slicenum)] = tmpFibrousIntegrity

                self.branchSliceAreaInfo['layer_' + str(slicenum)] = slicewallAreaInfo
                self.branchSliceAreaMaskInfo['layer_' + str(slicenum)] = slicewallAreaMaskInfo


                groupwallAreaInfo['layer_' + str(slicenum)] = slicewallAreaInfo
            self.wallAreaInfo['branch' + str(BGcont + 1)] = groupwallAreaInfo


            self.BranchGroupSliceLumenMaskAreadict['branch' + str(BGcont + 1)] = sliceLumenMaskArealist
            self.BranchGroupslicePlaqueAreaBooldict['branch' + str(BGcont + 1)] = slicePlaqueAreaBoollist

            self.branchGroupSliceAreaMaskInfo['branch' + str(BGcont + 1)] = self.branchSliceAreaMaskInfo
            self.branchGroupSliceAreaInfo['branch' + str(BGcont + 1)] = self.branchSliceAreaInfo

        print("BranchGroupslicePlaqueAreaBooldict", self.BranchGroupslicePlaqueAreaBooldict)
        print("branchGrpupSliceAreaMaskInfo", self.branchGroupSliceAreaMaskInfo)
        print("branchGrpupSliceInfo", self.branchGroupSliceAreaInfo)

        #print("self.BranchGroupslicePlaqueAreaBooldict", self.BranchGroupslicePlaqueAreaBooldict)

        self.TotalCalcificationArea = self.TotalCalcificationMaskArea * self.spacing_x * self.spacing_y
        self.TotalCalcificationVolumn = self.TotalCalcificationMaskArea * self.spacing_x * self.spacing_y * self.spacing_z
        try:
            self.TotalCalcificationVolumnRatio = self.TotalCalcificationMaskArea /self.TotalPlaqueMaskArea
        except:
            self.TotalCalcificationVolumnRatio = None
        print("CalcificationVolumn",self.TotalCalcificationVolumn)
        print("CalcificationVolumnRatio",self.TotalCalcificationVolumnRatio)

        #####
        self.TotalEnhancementArea = self.TotalEnhancementMaskArea * \
                                      self.spacing_x * self.spacing_y
        self.TotalEnhancementVolumn = self.TotalEnhancementMaskArea * \
                                        self.spacing_x * self.spacing_y * \
                                        self.spacing_z
        try:
            self.TotalEnhancementVolumnRatio = \
                self.TotalEnhancementMaskArea / self.TotalPlaqueMaskArea
        except:
            self.TotalEnhancementVolumnRatio = None
        print("EnhancementVolumn", self.TotalEnhancementVolumn)
        print("EnhancementVolumnRatio", self.TotalEnhancementVolumnRatio)

        self.TotalLipidArea = self.TotalLipidMaskArea * self.spacing_x * self.spacing_y
        self.TotalLipidVolumn = self.TotalLipidMaskArea * self.spacing_x * self.spacing_y * self.spacing_z
        try:
            self.TotalLipidVolumnRatio = self.TotalLipidMaskArea \
                                         / self.TotalPlaqueMaskArea
        except:
            self.TotalLipidVolumnRatio = None
        print("LipidVolumn", self.TotalLipidVolumn)
        print("LipidVolumnRatio", self.TotalLipidVolumnRatio)

        self.TotalHemorrhageArea = self.TotalHemorrhageMaskArea * self.spacing_x * self.spacing_y
        print("self.TotalHemorrhageMaskArea", self.TotalHemorrhageMaskArea)
        self.TotalHemorrhageVolumn = self.TotalHemorrhageMaskArea  * self.spacing_x * self.spacing_y * self.spacing_z
        try:
            self.TotalHemorrhageVolumnRatio = self.TotalHemorrhageMaskArea \
                                              / self.TotalPlaqueMaskArea
        except:
            self.TotalHemorrhageVolumnRatio = None
        print("HemorrhageVolumn", self.TotalHemorrhageVolumn)
        print("HemorrhageVolumnRatio", self.TotalHemorrhageVolumnRatio)

        self.TotalFibrousArea = self.TotalFibrousMaskArea * self.spacing_x * self.spacing_y
        self.TotalFibrousVolumn = self.TotalFibrousMaskArea  * self.spacing_x * self.spacing_y * self.spacing_z
        try:
            self.TotalFibrousVolumnRatio = self.TotalFibrousMaskArea \
                                           / self.TotalPlaqueMaskArea
        except:
            self.TotalFibrousVolumnRatio = None
        print("FibrousVolumn", self.TotalFibrousVolumn)
        print("FibrousVolumnRatio", self.TotalFibrousVolumnRatio)

        self.TotalPlaqueArea = self.TotalPlaqueMaskArea * self.spacing_x * self.spacing_y
        self.TotalPlaqueArea_junzong = self.TotalPlaqueMaskArea_junzong * self.spacing_x * self.spacing_y
        self.TotalPlaqueVolumn = self.TotalPlaqueMaskArea * self.spacing_x * self.spacing_y * self.spacing_z
        self.TotalPlaqueVolumn_junzong = self.TotalPlaqueMaskArea_junzong * self.spacing_x * self.spacing_y * self.spacing_z
        print("PlaqueVolumn", self.TotalPlaqueVolumn)

        self.TotalLumenArea = self.TotalLumenMaskArea * self.spacing_x * self.spacing_y
        self.AverLumenArea = self.TotalLumenArea / self.slicesNum
        print("AverLumenArea", self.AverLumenArea)

        AlllumenMaskArealist =[]
        for BGcont in range(len(self.BranchGroup)):
            AlllumenMaskArealist += self.BranchGroupSliceLumenMaskAreadict['branch' + str(BGcont + 1)]
        AlllumenMaskArealist.sort(reverse=False)#asceding order
        self.LumenAreaRange.append(AlllumenMaskArealist[0] * self.spacing_x * self.spacing_y)
        self.LumenAreaRange.append(AlllumenMaskArealist[-1] * self.spacing_x * self.spacing_y)
        print("LumenAreaRange",self.LumenAreaRange)

    def CarotidStenosis(self):
        # ******************************************the degree of stenosis****************************************** maximal stenosis degree
        self.stenosisLumenEqDiameter = {}
        self.stenosisPositionInfo = {}
        self.BranchGroupreferenceWallEquivalentDiameterIndex = {}
        self.BranchGroupreferenceWallEquivalentDiameter = {}
        self.BranchGroupReconstructionIndex = {}
        self.BranchGroupStenosisDegree = {}
        self.ReconstructionIndexInfo = {}
        self.ReconstructionIndexInfoMin = {}
        ReconstructionIndexlist = []
        #####junzong
        ReconstructionIndexlistMin = []
        self.StenosisDegreeList = []
        referenceWallEquivalentDiameterIndex = 0
        for BGcont in range(len(self.BranchGroup)):
            try:
                # ********************************* minimum lumen equivalent diameter***********************************
                tmpsliceLumenEqDiameterlistOri= copy.deepcopy(self.LumenEqDiameterInfo['branch' + str(BGcont + 1)])
                tmpsliceLumenEqDiameterlistSort = copy.deepcopy(self.LumenEqDiameterInfo['branch' + str(BGcont + 1)])
                tmpsliceWallEqDiameterlist = self.WallEqDiameterInfo['branch' + str(BGcont + 1)]
                tmpslicePlaqueBoollist = self.BranchGroupslicePlaqueAreaBooldict['branch' + str(BGcont + 1)]
                #print("tmpslicePlaqueBoollist", tmpslicePlaqueBoollist)

                tmpSliceWallArealist = self.WallAreaInfo['branch' + str(BGcont + 1)]
                tmpSliceWallArea_atPlaque = numpy.multiply(tmpSliceWallArealist, tmpslicePlaqueBoollist)
                #print("tmpSliceWallArea_atPlaque", tmpSliceWallArea_atPlaque)

                #####junzong
                tmpSliceWallArea_atPlaque = list(filter(lambda a: a != 0, tmpSliceWallArea_atPlaque))
                tmpSliceWallArea_atPlaque.sort()
                print("#####################################tmpSliceWallArea_atPlaque",tmpSliceWallArea_atPlaque)
                tmpSliceWallArea_atPlaquelargest = tmpSliceWallArea_atPlaque[-1]
                tmpSliceWallArea_atPlaqueMin = tmpSliceWallArea_atPlaque[0]

                # ********************************* minimum lumen equivalent diameter***********************************
                if sum(tmpslicePlaqueBoollist):
                    for slicenum in range(len(self.BranchGroup['branch' + str(BGcont + 1)])):
                        if tmpslicePlaqueBoollist[slicenum] == 0:  # if the current slice has no plaque
                            tmpsliceLumenEqDiameterlistSort[slicenum] = pow(10, 5)
                    min_index = tmpsliceLumenEqDiameterlistSort.index(
                        min(tmpsliceLumenEqDiameterlistSort))
                    minimumLumenEqDiameter = tmpsliceLumenEqDiameterlistSort[
                        min_index]
                    self.stenosisLumenEqDiameter[
                        'branch' + str(BGcont + 1)] = minimumLumenEqDiameter
                    self.stenosisPositionInfo[
                        'branch' + str(BGcont + 1)] = min_index
                else:
                    min_index = tmpsliceLumenEqDiameterlistSort.index(
                        min(tmpsliceLumenEqDiameterlistSort))
                    minimumLumenEqDiameter = tmpsliceLumenEqDiameterlistSort[min_index]
                    self.stenosisPositionInfo[
                        'branch' + str(BGcont + 1)] = min_index
                print("minimumLumenEqDiameter", minimumLumenEqDiameter)

                # ********************************* reference wall equivalent diameter**********************************
                if sum(tmpslicePlaqueBoollist) != 0:
                    PBoolcont = min_index
                    Goforward = 1
                    GoBack = 1
                    while PBoolcont <= len(self.BranchGroup['branch' + str(BGcont + 1)])-1 and Goforward ==1: #go forward
                        if tmpslicePlaqueBoollist[PBoolcont] != 0:
                            PBoolcont += 1
                        else:
                            referenceWallEquivalentDiameterIndex = PBoolcont
                            Goforward = 0
                    if Goforward == 1:#  go back
                        PBoolcont = min_index
                        while PBoolcont >= 0 and GoBack == 1:
                            if tmpslicePlaqueBoollist[PBoolcont] != 0:
                                PBoolcont -= 1
                            else:
                                referenceWallEquivalentDiameterIndex = PBoolcont
                                GoBack = 0
                    #print("GoBack", GoBack)
                    self.BranchGroupreferenceWallEquivalentDiameterIndex[
                        'branch' + str(
                            BGcont + 1)] = referenceWallEquivalentDiameterIndex

                    self.BranchGroupreferenceWallEquivalentDiameter[
                        'branch' + str(BGcont + 1)] = \
                        tmpsliceWallEqDiameterlist[
                            referenceWallEquivalentDiameterIndex]
                    self.StenosisDegreeList.append(1 - numpy.square(minimumLumenEqDiameter /
                                                   tmpsliceLumenEqDiameterlistOri[
                                                       referenceWallEquivalentDiameterIndex]))
                    tmpStenosisDegreeList = (
                        1 - numpy.square(minimumLumenEqDiameter /
                                         tmpsliceLumenEqDiameterlistOri[
                                             referenceWallEquivalentDiameterIndex]))
                else:
                    self.StenosisDegreeList.append(1 - numpy.square(minimumLumenEqDiameter / numpy.mean(tmpsliceLumenEqDiameterlistOri)))
                    tmpStenosisDegreeList = (1 - numpy.square(minimumLumenEqDiameter / numpy.mean(tmpsliceLumenEqDiameterlistOri)))


                self.BranchGroupStenosisDegree['branch' + str(BGcont +
                                                              1)] = tmpStenosisDegreeList

                ReconstructionIndexlist.append(
                    tmpSliceWallArea_atPlaquelargest / tmpSliceWallArealist[
                        referenceWallEquivalentDiameterIndex])
                self.BranchGroupReconstructionIndex['branch' + str(
                    BGcont + 1)] = tmpSliceWallArea_atPlaquelargest/tmpSliceWallArealist[referenceWallEquivalentDiameterIndex]

                #####junzong
                ReconstructionIndexlistMin.append(
                    tmpSliceWallArea_atPlaqueMin / tmpSliceWallArealist[
                        referenceWallEquivalentDiameterIndex])
                self.BranchGroupReconstructionIndex['branch' + str(
                    BGcont + 1)] = tmpSliceWallArea_atPlaqueMin / \
                                   tmpSliceWallArealist[
                                       referenceWallEquivalentDiameterIndex]

                self.ReconstructionIndexInfo['branch' + str(BGcont + 1)] = tmpSliceWallArea_atPlaquelargest / tmpSliceWallArealist[
                        referenceWallEquivalentDiameterIndex]
                #####junzong
                self.ReconstructionIndexInfoMin['branch' + str(
                    BGcont + 1)] = tmpSliceWallArea_atPlaqueMin / \
                                   tmpSliceWallArealist[
                                       referenceWallEquivalentDiameterIndex]
            except:
                self.WebrtrnInfo["warningMessage"] += ": branch " + str(
                    BGcont) + " is blank"
                self.JsonFileOutput()

        self.StenosisDegreeList.sort()
        print("self.StenosisDegreeList", self.StenosisDegreeList)
        self.stenosisdegree = self.StenosisDegreeList[-1]
        ReconstructionIndexlist.sort()
        print("ReconstructionIndexlist",ReconstructionIndexlist)
        print("self.ReconstructionIndexInfo",self.ReconstructionIndexInfo)
        #####junzong
        ReconstructionIndexlistMin.sort()
        print("ReconstructionIndexlistMin", ReconstructionIndexlistMin)
        print("self.ReconstructionIndexInfoMin", self.ReconstructionIndexInfoMin)


        self.ReconstructionIndex = ReconstructionIndexlist[-1]
        self.ReconstructionIndexMin = ReconstructionIndexlistMin[0]   #the minimum value of all branches

        print("stenosisLumenEqDiameter: ", self.stenosisLumenEqDiameter)
        print("stenosisPositionInfo: ", self.stenosisPositionInfo)
        print("BranchGroupreferenceWallEquivalentDiameterIndex: ", self.BranchGroupreferenceWallEquivalentDiameterIndex)
        print("BranchGroupreferenceWallEquivalentDiameter: ", self.BranchGroupreferenceWallEquivalentDiameter)
        print("BranchGroupStenosisDegree: ", self.BranchGroupStenosisDegree)

        print("stenosisdegree: ", self.stenosisdegree)


        print("ReconstructionIndex: ", self.ReconstructionIndex)
        print("ReconstructionIndexMin: ", self.ReconstructionIndexMin)


    def wallProperty(self):
        self.wallthicknessInfo = {}
        self.WallThicknessRange = []
        wallmaskthicknesslist = []
        walleccentricitylist = []
        try:
            for BGcont in range(len(self.BranchGroup)):
                slicewallthickness = {}
                for slicenum in range(len(self.BranchGroup['branch' + str(BGcont + 1)])):
                    slicewallintersectpointsdict = self.sliceIntersectPointsInfo[
                        'branch_' + str(BGcont + 1) + "/layer_" + str(slicenum)]

                    contourintersectpointsinfoA = slicewallintersectpointsdict[0]
                    contourintersectpointsinfoB = slicewallintersectpointsdict[1]

                    singleAnglewallthickness = {}
                    singleAnglewalleccentricity = []
                    for pointcont in range(len(contourintersectpointsinfoA)):  # intersect points num of each angel is 2
                        pa_1 = contourintersectpointsinfoA[pointcont][0]
                        pa_2 = contourintersectpointsinfoA[pointcont][1]
                        pb_1 = contourintersectpointsinfoB[pointcont][0]
                        pb_2 = contourintersectpointsinfoB[pointcont][1]

                        # singleAngleWallthicknessA = numpy.sqrt(
                        #     numpy.square(abs(pb_1[0] - pa_1[0]) +1) + numpy.square(abs(pb_1[1] - pa_1[1]) +1))
                        singleAngleWallthicknessA = numpy.sqrt(
                            numpy.square(abs(pb_1[0] - pa_1[0])) + numpy.square(abs(pb_1[1] - pa_1[1]))) #  the intersection points are close to the boundary in opencv
                        if singleAngleWallthicknessA ==0:
                            singleAngleWallthicknessA = 1

                        singleAngleWallthicknessB = numpy.sqrt(
                            numpy.square(abs(pb_2[0] - pa_2[0])) + numpy.square(abs(pb_2[1] - pa_2[1])))
                        if singleAngleWallthicknessB ==0:
                            singleAngleWallthicknessB = 1

                        singleAnglewallthickness["angle_" + str(pointcont * 360 / len(contourintersectpointsinfoA))] = [
                            singleAngleWallthicknessA, singleAngleWallthicknessB]
                        wallmaskthicknesslist.append(singleAngleWallthicknessA)
                        wallmaskthicknesslist.append(singleAngleWallthicknessB)

                        walleccentricity = numpy.maximum(singleAngleWallthicknessA, singleAngleWallthicknessB)/numpy.minimum(singleAngleWallthicknessA, singleAngleWallthicknessB)
                        #singleAnglewalleccentricity["angle_" + str(pointcont * 360 / len(contourintersectpointsinfoA))] = walleccentricity
                        walleccentricitylist.append(walleccentricity)

                    slicewallthickness['layer_' + str(slicenum)] = singleAnglewallthickness
                self.wallthicknessInfo['branch' + str(BGcont + 1)] = slicewallthickness
        except:
            self.WebrtrnInfo["warningMessage"] += "There is no wall!"
            self.JsonFileOutput()
            # sys.exit(0)

        wallmaskthicknesslist.sort()
        walleccentricitylist.sort()
        print("wallmaskthicknesslist", wallmaskthicknesslist)

        self.AverWallThickness = sum(wallmaskthicknesslist) / len(wallmaskthicknesslist) * self.spacing_x
        self.WallThicknessRange.append(wallmaskthicknesslist[0] * self.spacing_x)
        self.WallThicknessRange.append(wallmaskthicknesslist[-1] * self.spacing_x)
        self.walleccentricity = walleccentricitylist[-1]
        print("wallthicknessInfo", self.wallthicknessInfo)


        print("AverWallThickness", self.AverWallThickness)


        print("WallThicknessRange", self.WallThicknessRange)


        print("walleccentricity", self.walleccentricity)

##########################################################################################################

    def JsonFileOutput(self):
        json_str = json.dumps(self.PlaqueAnalysisInfo)
        with open(self.OutputFilePath, 'w') as json_file:
        #with open(r'E:\tst.json', 'w') as json_file:
            json_file.write(json_str)


    def xlsxFileOutput(self):
        wb = openpyxl.Workbook()
        # grab the active worksheet
        ws = wb.active
        # Data can be assigned directly to cells and python types will automatically be converted
        # ws['A1'] = 42
        # Rows can also be appended
        ws.append(['spacing_x', 'spacing_y', 'spacing_z',
                   'PlaqueBurden','PlaqueHeight','CalcificationVolumn',
                   'LipidVolumn','HemorrhageVolumn', 'FibrousVolumn',
                   'EHVolumn','PlaqueVolumn','PlaqueVolumn_junzong','stenosisdegree',
                   'ReconstructionIndexMax','ReconstructionIndexMin','AverWallThickness','WallThicknessRange',
                   'walleccentricity','CalcificationVolumnRatio','LipidVolumnRatio',
                   'HemorrhageVolumnRatio', 'FibrousVolumnRatio','EHVolumnRatio'])
        ws.append([self.spacing_x, self.spacing_y, self.spacing_z,
                   self.PlaqueBurden, self.PlaqueHeight, self.TotalCalcificationVolumn,
                   self.TotalLipidVolumn, self.TotalHemorrhageVolumn, self.TotalFibrousVolumn,
                   self.TotalEnhancementVolumn, self.TotalPlaqueVolumn, self.TotalPlaqueVolumn_junzong, self.stenosisdegree,
                   self.ReconstructionIndex, self.ReconstructionIndexMin,self.AverWallThickness,str(self.WallThicknessRange),
                   self.walleccentricity, self.TotalCalcificationVolumnRatio,self.TotalLipidVolumnRatio,
                   self.TotalHemorrhageVolumnRatio, self.TotalFibrousVolumnRatio,self.TotalEnhancementVolumnRatio])
        ws.append([""])
        ws.append(["SliceAreaMaskInfo"]) #self.PlaqueAnalysisInfo

        xlsx_SliceAreaMaskInfo = self.PlaqueAnalysisInfo["SliceAreaMaskInfo"]
        for count in range(len(xlsx_SliceAreaMaskInfo)):
            ws.append(['branch'+str(count + 1), 'Lumen', 'Wall', 'Calcification',	'Lipid', 'Hemorrhage', 'Fibrous', 'EH'])
            xlsx_SliceAreaMaskInfo_brach = xlsx_SliceAreaMaskInfo['branch' + str(count + 1)]
            for layerCount in range(len(xlsx_SliceAreaMaskInfo_brach)):
                xlsx_SliceAreaMaskInfo_brach_layer = xlsx_SliceAreaMaskInfo_brach['layer_' + str(layerCount)]
                ws.append(['layer' + str(layerCount + 1),
                           xlsx_SliceAreaMaskInfo_brach_layer['Lumen'],
                           xlsx_SliceAreaMaskInfo_brach_layer['Wall'],
                           xlsx_SliceAreaMaskInfo_brach_layer['Calcification'],
                           xlsx_SliceAreaMaskInfo_brach_layer['Lipid'],
                           xlsx_SliceAreaMaskInfo_brach_layer['Hemorrhage'],
                           xlsx_SliceAreaMaskInfo_brach_layer['Fibrous'],
                           xlsx_SliceAreaMaskInfo_brach_layer['Enhancement']
                           ])
        ws.append([""])
        ws.append(["SliceAreaInfo"])  # self.PlaqueAnalysisInfo
        xlsx_SliceAreaInfo = self.PlaqueAnalysisInfo["SliceAreaInfo"]
        for count in range(len(xlsx_SliceAreaInfo)):
            ws.append(['branch' + str(count + 1), 'Lumen', 'Wall',
                       'Calcification', 'Lipid', 'Hemorrhage', 'Fibrous',
                       'Enhancement'])
            xlsx_SliceAreaInfo_brach = xlsx_SliceAreaInfo['branch' + str(count + 1)]
            for layerCount in range(len(xlsx_SliceAreaInfo_brach)):
                xlsx_SliceAreaInfo_brach_layer = \
                xlsx_SliceAreaInfo_brach['layer_' + str(layerCount)]
                ws.append(['layer' + str(layerCount + 1),
                           xlsx_SliceAreaInfo_brach_layer['Lumen'],
                           xlsx_SliceAreaInfo_brach_layer['Wall'],
                           xlsx_SliceAreaInfo_brach_layer[
                               'Calcification'],
                           xlsx_SliceAreaInfo_brach_layer['Lipid'],
                           xlsx_SliceAreaInfo_brach_layer['Hemorrhage'],
                           xlsx_SliceAreaInfo_brach_layer['Fibrous'],
                           xlsx_SliceAreaInfo_brach_layer['Enhancement']
                           ])



        # ws.cell(row=2, column=1).value = 'spacing_x'
        # ws.cell(row=2, column=2).value = 'spacing_y'
        # ws.cell(row=2, column=3).value = 'spacing_z'
        # ws.cell(row=2, column=4).value = 'PlaqueBurden'
        # ws.cell(row=2, column=5).value = 'PlaqueHeight'
        # ws.cell(row=2, column=6).value = 'CalcificationVolumn'
        # ws.cell(row=2, column=7).value = 'LipidVolumn'
        # ws.cell(row=2, column=8).value = 'HemorrhageVolumn'
        # ws.cell(row=2, column=9).value = 'FibrousVolumn'
        # ws.cell(row=2, column=10).value = 'EHVolumn'
        # ws.cell(row=2, column=11).value = 'PlaqueVolumn'
        # ws.cell(row=2, column=12).value = 'stenosisdegree'
        # ws.cell(row=2, column=13).value = 'ReconstructionIndex'
        # ws.cell(row=2, column=14).value = 'AverWallThickness'
        # ws.cell(row=2, column=15).value = 'WallThicknessRange'
        # ws.cell(row=2, column=16).value = 'walleccentricity'
        # ws.cell(row=2, column=17).value = 'CalcificationVolumnRatio'
        # ws.cell(row=2, column=18).value = 'LipidVolumnRatio'
        # ws.cell(row=2, column=19).value = 'HemorrhageVolumnRatio'


        #

        # Save the file
        wb.save(self.xlsxSavePath)


    def carotidplt(self):
        #pass self.BranchSliceGroup_ResetLabelDict['layer' + str(slicenum + 1)]
        for slicenum in range(0, len(self.BranchGroup['branch3'])):
            BranchSliceofall = self.BranchGroup['branch3'][slicenum]
            plt.figure()
            plt.imshow(BranchSliceofall)
        plt.show()


#
# work_dir = sys.argv[1]
# output_dir = sys.argv[2]
# print ("##### sys.argv #######",sys.argv)


#work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\luyu\cal4_1.3.46.670589.11.34312 .5.0.10496.2015080519255095584_left.nii.gz"
# work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\luyu\1105\cal4_1.2.392.200036.9125.2.138612190166.20211014002433_left.nii.gz"
#work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\luyu\1103cal4_1.2.840.113820.104.22395.120211014144335970_right.nii.gz"
#work_dir = r"G:\Desktop\a11mask\mask.nii.gz"
#work_dir = r"G:\Desktop\toLuyu\mask.nii.gz"
#work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\luyu\output\testly_1.3.12.2.1107.5.2.19.45509.2021090614044381143077871.0.0.0L_Carotid.nii.gz"
#work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\luyu\output\01334157\R_Carotid.nii.gz"
#work_dir = r"E:\Carotid\nii\R_plaque_classify_8class_decrop2.nii.gz"
#work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\luoming\cal19_1.3.12.2.1107.5.2.20.156152.30000021081407250006400000073_right.nii.gz"
# work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\siyuan\L_Carotid.nii.gz"
# work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\luyu\0107\cal52_1.2.840.113564.345051829620.1116348.637711166325284975.268_left.nii.gz"
# work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\YDH_Junzong\t1_tse_tra_0.5X0.5X2_R.nii.gz"
# work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\YDH_Junzong\OAx QIR fs T11.nii.gz"
# work_dir = r"E:\B_PlaqueQuantitativeAnalysis\Carotid\YDH_Junzong\new_t2_tse_tra_0.5X0.5X2_R.nii.gz"
# output_dir = r"E:\tstz.json"
# xlsxOutput_dir = r"E:\tstz.xlsx"


# object_dir = r"E:\B_PlaqueQuantitativeAnalysis\cerebral_artery\output0117\output"
object_dir = r"E:\A_Projects\A_Hemodynamics_Platform\Hedys\GUI_V0\Modules_YC\PlaqueQuantitativeAnalysis\output0304"
# for root, dirs, files in os.walk(object_dir):
#     if 'CTA.nii.gz' in files:
#         newPath = root.replace('RawData', 'OriBranch')
#         cmd = 'copy ' + os.path.join(root, 'CTA.nii.gz') + ' ' + os.path.join(newPath, 'CTA.nii.gz')
#         print(cmd)
#         # os.system(cmd)
for root, dirs, files in os.walk(object_dir):
    if files:
        for file in files:
            if file.endswith('.nii.gz') and file.startswith('new_'):
                try:
                    work_dir = os.path.join(root,file)
                    xlsxName = file.replace('.nii.gz', '.xlsx')
                    jsonName = file.replace('.nii.gz', '.json')
                    xlsxOutput_dir = os.path.join(root,xlsxName)
                    jsonOutput_dir = os.path.join(root,xlsxName)
                    # PQA = PlaqueQuantitativAnalysis(work_dir, output_dir)
                    PQA = PlaqueQuantitativAnalysis(work_dir, jsonOutput_dir,
                                                    xlsxOutput_dir)
                    PQA.PlaqueAnalysisInfo = {}
                    PQA.LoadData()
                    PQA.DataBifurcationRemove()
                    ######################################## Normal ################################
                    PQA.outputData_cont = copy.deepcopy(PQA.outputData)
                    PQA.LumenBinary()
                    PQA.NonBlanklayerIdGet()
                    PQA.BifurcationGet()
                    PQA.BranchSplited()
                    PQA.BranchGrouping()
                    # PQA.carotidplt()
                    PQA.IntersectPoints()
                    PQA.CarotidTissueParameters()
                    PQA.CarotidStenosis()
                    PQA.wallProperty()

                    if PQA.WebrtrnInfo["warningMessage"]:
                        PQA.PlaqueAnalysisInfo[
                            "warningMessage"] = "First step with bifurcation: " + \
                                                PQA.WebrtrnInfo["warningMessage"]
                    else:
                        PQA.PlaqueAnalysisInfo["warningMessage"] = PQA.WebrtrnInfo[
                            "warningMessage"]

                    if PQA.WebrtrnInfo["errorMessage"]:
                        PQA.PlaqueAnalysisInfo[
                            "errorMessage"] = "First step with bifurcationL: " + \
                                              PQA.WebrtrnInfo["errorMessage"]
                    else:
                        PQA.PlaqueAnalysisInfo["errorMessage"] = PQA.WebrtrnInfo[
                            "errorMessage"]

                    PQA.PlaqueAnalysisInfo["log"] = "First step finished! " + \
                                                    PQA.WebrtrnInfo["log"]

                    PQA.PlaqueAnalysisInfo["PlaqueBurden"] = PQA.PlaqueBurden

                    PQA.PlaqueAnalysisInfo["PlaqueHeight"] = PQA.PlaqueHeight

                    # PQA.PlaqueAnalysisInfo["FibrousIntegrity"] = PQA.BranchGroupFibrousIntegrity

                    PQA.PlaqueAnalysisInfo[
                        "SliceAreaMaskInfo"] = PQA.branchGroupSliceAreaMaskInfo
                    PQA.PlaqueAnalysisInfo[
                        "SliceAreaInfo"] = PQA.branchGroupSliceAreaInfo

                    PQA.PlaqueAnalysisInfo[
                        "CalcificationVolumn"] = PQA.TotalCalcificationVolumn
                    PQA.PlaqueAnalysisInfo[
                        "CalcificationVolumnRatio"] = PQA.TotalCalcificationVolumnRatio

                    PQA.PlaqueAnalysisInfo[
                        "EnhancementVolumn"] = PQA.TotalEnhancementVolumn
                    PQA.PlaqueAnalysisInfo[
                        "EnhancementVolumnRatio"] = PQA.TotalEnhancementVolumnRatio

                    PQA.PlaqueAnalysisInfo["LipidVolumn"] = PQA.TotalLipidVolumn
                    PQA.PlaqueAnalysisInfo[
                        "LipidVolumnRatio"] = PQA.TotalLipidVolumnRatio

                    PQA.PlaqueAnalysisInfo[
                        "HemorrhageVolumn"] = PQA.TotalHemorrhageVolumn
                    PQA.PlaqueAnalysisInfo[
                        "HemorrhageVolumnRatio"] = PQA.TotalHemorrhageVolumnRatio

                    PQA.PlaqueAnalysisInfo["FibrousVolumn"] = PQA.TotalFibrousVolumn
                    PQA.PlaqueAnalysisInfo[
                        "FibrousVolumnRatio"] = PQA.TotalFibrousVolumnRatio

                    PQA.PlaqueAnalysisInfo["PlaqueVolumn"] = PQA.TotalPlaqueVolumn

                    PQA.PlaqueAnalysisInfo["AverLumenArea"] = PQA.AverLumenArea

                    PQA.PlaqueAnalysisInfo["LumenAreaRange"] = PQA.LumenAreaRange

                    PQA.PlaqueAnalysisInfo["wallAreaInfo"] = PQA.wallAreaInfo

                    PQA.PlaqueAnalysisInfo["stenosisdegree"] = PQA.stenosisdegree

                    PQA.PlaqueAnalysisInfo[
                        "ReconstructionIndex"] = PQA.ReconstructionIndex  # do
                    # not count the reconstruction index at the bifurcation
                    PQA.PlaqueAnalysisInfo[
                        "ReconstructionIndexInfo"] = PQA.ReconstructionIndexInfo

                    PQA.PlaqueAnalysisInfo[
                        "AverWallThickness"] = PQA.AverWallThickness

                    PQA.PlaqueAnalysisInfo[
                        "WallThicknessRange"] = PQA.WallThicknessRange

                    PQA.PlaqueAnalysisInfo[
                        "walleccentricity"] = PQA.walleccentricity
                    PQA.JsonFileOutput()
                    PQA.xlsxFileOutput()
                except:
                    print("error")
print('done')

# #PQA = PlaqueQuantitativAnalysis(work_dir, output_dir)
# PQA = PlaqueQuantitativAnalysis( work_dir, output_dir, xlsxOutput_dir)
# PQA.PlaqueAnalysisInfo ={}
# PQA.LoadData()
# PQA.DataBifurcationRemove()
# ######################################## Normal ################################
# PQA.outputData_cont = copy.deepcopy(PQA.outputData)
# PQA.LumenBinary()
# PQA.NonBlanklayerIdGet()
# PQA.BifurcationGet()
# PQA.BranchSplited()
# PQA.BranchGrouping()
# #PQA.carotidplt()
# PQA.IntersectPoints()
# PQA.CarotidTissueParameters()
# PQA.CarotidStenosis()
# PQA.wallProperty()
#
# if PQA.WebrtrnInfo["warningMessage"]:
#     PQA.PlaqueAnalysisInfo["warningMessage"] = "First step with bifurcation: " + \
#                                       PQA.WebrtrnInfo["warningMessage"]
# else:
#     PQA.PlaqueAnalysisInfo["warningMessage"] = PQA.WebrtrnInfo["warningMessage"]
#
# if PQA.WebrtrnInfo["errorMessage"]:
#     PQA.PlaqueAnalysisInfo["errorMessage"] = "First step with bifurcationL: " + \
#                           PQA.WebrtrnInfo["errorMessage"]
# else:
#     PQA.PlaqueAnalysisInfo["errorMessage"] = PQA.WebrtrnInfo["errorMessage"]
#
# PQA.PlaqueAnalysisInfo["log"] = "First step finished! " + \
#                             PQA.WebrtrnInfo["log"]
#
# PQA.PlaqueAnalysisInfo["PlaqueBurden"]= PQA.PlaqueBurden
#
# PQA.PlaqueAnalysisInfo["PlaqueHeight"] = PQA.PlaqueHeight
#
# # PQA.PlaqueAnalysisInfo["FibrousIntegrity"] = PQA.BranchGroupFibrousIntegrity
#
# PQA.PlaqueAnalysisInfo["SliceAreaMaskInfo"] = PQA.branchGroupSliceAreaMaskInfo
# PQA.PlaqueAnalysisInfo["SliceAreaInfo"] = PQA.branchGroupSliceAreaInfo
#
# PQA.PlaqueAnalysisInfo["CalcificationVolumn"] = PQA.TotalCalcificationVolumn
# PQA.PlaqueAnalysisInfo["CalcificationVolumnRatio"] = PQA.TotalCalcificationVolumnRatio
#
# PQA.PlaqueAnalysisInfo["EnhancementVolumn"] = PQA.TotalEnhancementVolumn
# PQA.PlaqueAnalysisInfo["EnhancementVolumnRatio"] = PQA.TotalEnhancementVolumnRatio
#
# PQA.PlaqueAnalysisInfo["LipidVolumn"] = PQA.TotalLipidVolumn
# PQA.PlaqueAnalysisInfo["LipidVolumnRatio"] = PQA.TotalLipidVolumnRatio
#
# PQA.PlaqueAnalysisInfo["HemorrhageVolumn"] = PQA.TotalHemorrhageVolumn
# PQA.PlaqueAnalysisInfo["HemorrhageVolumnRatio"] = PQA.TotalHemorrhageVolumnRatio
#
# PQA.PlaqueAnalysisInfo["FibrousVolumn"] = PQA.TotalFibrousVolumn
# PQA.PlaqueAnalysisInfo["FibrousVolumnRatio"] = PQA.TotalFibrousVolumnRatio
#
# PQA.PlaqueAnalysisInfo["PlaqueVolumn"] = PQA.TotalPlaqueVolumn
#
# PQA.PlaqueAnalysisInfo["AverLumenArea"] = PQA.AverLumenArea
#
# PQA.PlaqueAnalysisInfo["LumenAreaRange"] = PQA.LumenAreaRange
#
# PQA.PlaqueAnalysisInfo["wallAreaInfo"] = PQA.wallAreaInfo
#
# PQA.PlaqueAnalysisInfo["stenosisdegree"] = PQA.stenosisdegree
#
# PQA.PlaqueAnalysisInfo["ReconstructionIndex"] = PQA.ReconstructionIndex  #do not count the reconstruction index at the bifurcation
# PQA.PlaqueAnalysisInfo["ReconstructionIndexInfo"] = PQA.ReconstructionIndexInfo
#
# PQA.PlaqueAnalysisInfo["AverWallThickness"] = PQA.AverWallThickness
#
# PQA.PlaqueAnalysisInfo["WallThicknessRange"] = PQA.WallThicknessRange
#
# PQA.PlaqueAnalysisInfo["walleccentricity"] = PQA.walleccentricity
# PQA.JsonFileOutput()
# PQA.xlsxFileOutput()
# print("PlaqueAnalysisInfo:", PQA.PlaqueAnalysisInfo)

# orilist = [0,0,1,1,0,0,0,1,0,1,0,1,1,1,1,0]
# listRemoveDuplicates = listRemoveDuplicates(orilist)
# branchNums = sum(listRemoveDuplicates)
#
# name = locals()
# for count in range(branchNums):
#     name["branch" + str(count)] = []
#
# # branch4= []
# # branch5= []
# # branch6= []
# # branch7= []
# # branch8= []
# # branch9= []
# # branch10= []
# listNum = 0
# for count in range(1, len(orilist)):
#     print(orilist[count])
#     name["branch" + str(listNum)].append(orilist[count -1])
#     if orilist[count] == orilist[count -1]:
#         name["branch" + str(listNum)].append(orilist[count])
#
#     else:
#         listNum += 1
#         print("current listNum", listNum)
#     print("branch" + str(listNum), name["branch" + str(listNum)])




